import pandas as pd
from openpyxl import load_workbook  # Importa load_workbook desde openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

def format_and_save_to_excel(data, file_name, sheet_name, TableStyle, column_widths=None):
    """
    Aplica el formato científico y ajusta el ancho de las columnas antes de guardar el DataFrame en un archivo Excel.
    """
    # Especificar el formato científico para todas las columnas
    formato_cientifico = "{:.4e}"  # Puedes ajustar el número de decimales según tus preferencias

    # Aplicar el formato científico a todas las columnas
    for columna in data.columns:
        data[columna] = data[columna].apply(lambda x: formato_cientifico.format(x))

    # Si el archivo no existe, crea un nuevo archivo Excel
    if not os.path.exists(file_name):
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            data.to_excel(writer, index=False, sheet_name=sheet_name)
    else:
        # Agrega el DataFrame a un archivo Excel existente
        with pd.ExcelWriter(file_name, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            data.to_excel(writer, index=False, sheet_name=sheet_name)

    if column_widths:
        # Abrir el archivo Excel y ajustar el ancho de las columnas
        wb = load_workbook(file_name)  # Usa load_workbook de openpyxl
        ws = wb[sheet_name]
        for i, column_width in enumerate(column_widths):
            ws.column_dimensions[ws.cell(row=1, column=i+1).column_letter].width = column_width

        # Centrar los datos en todas las celdas
        for row in ws.iter_rows(min_row=2, min_col=1):  # Excluye la fila de encabezados
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

    sheet_name_comprimido = sheet_name.replace(" ", "")
    # Aplicar formato de tabla
    table = Table(displayName=f"Table_{sheet_name_comprimido}", ref=ws.dimensions)
    style = TableStyleInfo(name=TableStyle, showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)

    wb.save(file_name)